from openpyxl import Workbook
from openpyxl import load_workbook


def main():
    workbook = load_workbook(filename="import.xlsx")
    workbook.sheetnames
    sheet = workbook.active
    print(sheet["A1"].value)
    # for row in sheet:
    #     print("%s: %s" % (row[0], row[1]))
    #     break


if __name__ == "__main__":
    main()
